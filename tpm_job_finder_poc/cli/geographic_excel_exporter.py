"""Enhanced Excel export system with geographic organization and multi-resume intelligence."""

import logging
from typing import Dict, List, Optional, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from tpm_job_finder_poc.enrichment.geographic_classifier import GeographicClassifier


logger = logging.getLogger(__name__)


class GeographicExcelExporter:
    """Excel exporter with geographic organization and optional multi-resume intelligence."""
    
    def __init__(self):
        """Initialize the geographic Excel exporter."""
        self.classifier = GeographicClassifier()
        
        # Style configurations
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.intelligence_font = Font(bold=True, size=10, italic=True)
        
        # Define the column structure according to specification
        self.ORIGINAL_COLUMNS = [
            'Title', 'Company', 'Location', 'Match Score', 
            'USD Salary', 'Visa Required', 'Source', 'Posted Date'
        ]
        
        # NEW: Specification-compliant column structure (A-L)
        self.SPEC_COLUMNS = [
            'Job Title',            # Column A (Keep - was 'Title')
            'Company',              # Column B (Keep) 
            'Location',             # Column C (Keep)
            'Salary',               # Column D (Keep - was 'USD Salary')
            'Selected Resume',      # Column E (NEW)
            'Match Score',          # Column F (Enhanced)
            'Selection Rationale',  # Column G (NEW)
            'Enhancement 1',        # Column H (NEW)
            'Enhancement 2',        # Column I (NEW)
            'Enhancement 3',        # Column J (NEW)
            'Job Source',           # Column K (Keep - was 'Source')
            'Date Posted',          # Column L (Keep - was 'Posted Date')
        ]
        
        # Style configurations
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.intelligence_font = Font(bold=True, size=10, italic=True)
        
        # Regional color scheme
        self.regional_colors = {
            'North America': 'E6F3FF',  # Light blue
            'Western Europe': 'E6FFE6',  # Light green
            'Eastern Europe': 'FFE6F3',  # Light pink
            'East Asia': 'FFF0E6',  # Light orange
            'Southeast Asia': 'F0E6FF',  # Light purple
            'South Asia': 'FFFFE6',  # Light yellow
            'South America': 'E6FFFF',  # Light cyan
            'Africa': 'F5E6D3',  # Light brown
            'Australia/Oceania': 'E6F0FF',  # Light lavender
            'Central America': 'FFE6E6',  # Light red
            'Other': 'F0F0F0'  # Light gray
        }
    
    def create_regional_workbook(self, jobs: List[Dict[str, Any]]) -> Workbook:
        """Create Excel workbook organized by geographic regions.
        
        Args:
            jobs: List of job dictionaries
            
        Returns:
            Excel workbook with regional tabs
        """
        return self._create_workbook(jobs, intelligence_results=None)
    
    def create_enhanced_regional_workbook(self, 
                                        jobs: List[Dict[str, Any]], 
                                        intelligence_results: List) -> Workbook:
        """Create enhanced Excel workbook with multi-resume intelligence.
        
        Args:
            jobs: List of job dictionaries
            intelligence_results: List of JobIntelligenceResult objects
            
        Returns:
            Excel workbook with regional tabs and enhanced columns
        """
        return self._create_workbook(jobs, intelligence_results)
    
    def _create_workbook(self, 
                        jobs: List[Dict[str, Any]], 
                        intelligence_results: Optional[List] = None) -> Workbook:
        """Create Excel workbook with optional multi-resume intelligence.
        
        Args:
            jobs: List of job dictionaries
            intelligence_results: Optional list of JobIntelligenceResult objects
            
        Returns:
            Excel workbook with regional tabs
        """
        logger.info(f"Creating workbook with {len(jobs)} jobs, enhanced={intelligence_results is not None}")
        
        # Enhance jobs with intelligence if provided
        if intelligence_results:
            jobs = self._enhance_jobs_with_intelligence(jobs, intelligence_results)
        
        # Organize jobs by region
        regional_jobs = self.classifier.organize_jobs_by_region(jobs)
        
        # Create workbook
        workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in [ws.title for ws in workbook.worksheets]:
            workbook.remove(workbook['Sheet'])
        
        # Create summary worksheet
        self._create_summary_worksheet(workbook, regional_jobs, intelligence_results is not None)
        
        # Create regional worksheets
        for region in self.classifier.get_region_priority_order():
            if region in regional_jobs and regional_jobs[region]:
                self._create_regional_worksheet(workbook, region, regional_jobs[region], intelligence_results is not None)
        
        logger.info(f"Created workbook with {len(workbook.worksheets)} worksheets")
        return workbook
    
    def _enhance_jobs_with_intelligence(self, 
                                      jobs: List[Dict[str, Any]], 
                                      intelligence_results: List) -> List[Dict[str, Any]]:
        """Enhance job data with multi-resume intelligence.
        
        Args:
            jobs: Original job data
            intelligence_results: Multi-resume intelligence results
            
        Returns:
            Enhanced job data with intelligence columns
        """
        # Create mapping of job_id to intelligence result
        intelligence_map = {result.job_id: result for result in intelligence_results}
        
        enhanced_jobs = []
        for job in jobs:
            enhanced_job = job.copy()
            
            # Get intelligence result for this job
            job_id = job.get('id', job.get('Job ID', ''))
            intelligence = intelligence_map.get(job_id)
            
            if intelligence:
                # Add multi-resume intelligence data
                enhanced_job.update({
                    # Column E: Selected Resume
                    'selected_resume': intelligence.selected_resume.filename if intelligence.selected_resume else 'N/A',
                    # Column F: Enhanced Match Score (override original)
                    'enhanced_match_score': intelligence.match_score,
                    # Column G: Selection Rationale
                    'selection_rationale': intelligence.selection_rationale or 'Multi-resume processing completed',
                    # Columns H, I, J: Enhancements
                    'enhancement_1': intelligence.enhancements[0].bullet_point if len(intelligence.enhancements) > 0 else '',
                    'enhancement_2': intelligence.enhancements[1].bullet_point if len(intelligence.enhancements) > 1 else '',
                    'enhancement_3': intelligence.enhancements[2].bullet_point if len(intelligence.enhancements) > 2 else '',
                })
            else:
                # Add empty intelligence columns
                enhanced_job.update({
                    'selected_resume': 'Not Available',
                    'enhanced_match_score': enhanced_job.get('match_score', 0),
                    'selection_rationale': 'Multi-resume processing not available',
                    'enhancement_1': '',
                    'enhancement_2': '',
                    'enhancement_3': '',
                })
            
            enhanced_jobs.append(enhanced_job)
        
        return enhanced_jobs
    
    def _create_summary_worksheet(self, 
                                workbook: Workbook, 
                                regional_jobs: Dict[str, List[Dict]], 
                                enhanced: bool = False):
        """Create summary worksheet with regional overview.
        
        Args:
            workbook: Excel workbook
            regional_jobs: Dictionary of jobs organized by region
            enhanced: Whether this includes multi-resume intelligence
        """
        ws = workbook.create_sheet(title="📊 Summary", index=0)
        
        # Title
        title = 'TPM Job Search Results - Geographic Summary'
        if enhanced:
            title += ' (Enhanced with Multi-Resume Intelligence)'
        
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Generation timestamp
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A2'].font = Font(italic=True)
        
        # Summary statistics
        total_jobs = sum(len(jobs) for jobs in regional_jobs.values())
        ws['A4'] = 'Overall Statistics'
        ws['A4'].font = self.header_font
        
        ws['A5'] = f'Total Jobs Found: {total_jobs}'
        ws['A6'] = f'Regions Covered: {len(regional_jobs)}'
        ws['A7'] = f'International Opportunities: {total_jobs - len(regional_jobs.get("North America", []))}'
        
        # Regional breakdown table
        ws['A9'] = 'Regional Breakdown'
        ws['A9'].font = self.header_font
        
        headers = ['Region', 'Job Count', 'Percentage', 'Top Company', 'Avg USD Salary']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = self.header_font
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type="solid")
        
        # Regional data
        row = 11
        for region in self.classifier.get_region_priority_order():
            if region in regional_jobs:
                jobs = regional_jobs[region]
                job_count = len(jobs)
                percentage = (job_count / total_jobs * 100) if total_jobs > 0 else 0
                
                # Calculate statistics
                top_company = self._get_top_company(jobs)
                avg_salary = self._calculate_avg_salary(jobs)
                
                ws.cell(row=row, column=1, value=region)
                ws.cell(row=row, column=2, value=job_count)
                ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
                ws.cell(row=row, column=4, value=top_company)
                ws.cell(row=row, column=5, value=f"${avg_salary:,.0f}" if avg_salary else "N/A")
                
                # Color code by region
                region_color = self.regional_colors.get(region, 'FFFFFF')
                fill = PatternFill(start_color=region_color, end_color=region_color, fill_type="solid")
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = fill
                
                row += 1
        
        # Auto-adjust column widths safely
        self._adjust_column_widths_safe(ws)
    
    def _create_regional_worksheet(self, 
                                 workbook: Workbook, 
                                 region: str, 
                                 jobs: List[Dict], 
                                 enhanced: bool = False):
        """Create worksheet for specific region.
        
        Args:
            workbook: Excel workbook
            region: Region name
            jobs: List of jobs for this region
            enhanced: Whether to use enhanced column structure
        """
        # Create worksheet with emoji and region name
        emoji = self._get_region_emoji(region)
        ws = workbook.create_sheet(title=f"{emoji} {region}")
        
        # Regional header and intelligence
        self._add_regional_intelligence(ws, region, enhanced)
        
        # Job data table with appropriate column structure
        self._add_job_data_table(ws, jobs, start_row=8, enhanced=enhanced)
        
        # Apply styling
        self._apply_regional_styling(ws, region)
        
        # Auto-adjust column widths
        self._adjust_column_widths_safe(ws)
    
    def _add_regional_intelligence(self, ws, region: str, enhanced: bool = False):
        """Add regional intelligence section to worksheet.
        
        Args:
            ws: Excel worksheet
            region: Region name
            enhanced: Whether this includes multi-resume intelligence
        """
        metadata = self.classifier.get_regional_metadata(region)
        
        # Regional title
        title = f'{region} - Regional Intelligence'
        if enhanced:
            title += ' + Multi-Resume Intelligence'
        
        ws['A1'] = title
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Intelligence data
        intelligence_data = [
            ('Business Culture:', metadata.get('business_culture', 'N/A')),
            ('Primary Timezones:', metadata.get('timezone_range', 'N/A')),
            ('Major Tech Hubs:', ', '.join(metadata.get('major_tech_hubs', []))),
            ('Cost of Living Adj:', f"{metadata.get('avg_cost_of_living', 1.0):.1f}x"),
            ('Currency Stability:', metadata.get('currency_stability', 'N/A')),
        ]
        
        for row, (label, value) in enumerate(intelligence_data, 3):
            ws.cell(row=row, column=1, value=label).font = self.intelligence_font
            ws.cell(row=row, column=2, value=value)
    
    def _add_job_data_table(self, ws, jobs: List[Dict], start_row: int = 8, enhanced: bool = False):
        """Add job data table to worksheet.
        
        Args:
            ws: Excel worksheet
            jobs: List of job dictionaries
            start_row: Starting row for the table
            enhanced: Whether to use enhanced column structure
        """
        # Choose column structure based on mode
        if enhanced:
            headers = self.SPEC_COLUMNS
        else:
            headers = self.ORIGINAL_COLUMNS
        
        # Headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type="solid")
            
            # Highlight NEW columns in enhanced mode
            if enhanced and header in ['Selected Resume', 'Selection Rationale', 'Enhancement 1', 'Enhancement 2', 'Enhancement 3']:
                cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type="solid")
        
        # Job data
        for row, job in enumerate(jobs, start_row + 1):
            if enhanced:
                self._add_enhanced_job_row(ws, row, job)
            else:
                self._add_original_job_row(ws, row, job)
    
    def _add_enhanced_job_row(self, ws, row: int, job: Dict):
        """Add job row with enhanced column structure (A-L).
        
        Args:
            ws: Excel worksheet
            row: Row number
            job: Job dictionary
        """
        # Column A: Job Title
        ws.cell(row=row, column=1, value=job.get('title', 'N/A'))
        
        # Column B: Company  
        ws.cell(row=row, column=2, value=job.get('company', 'N/A'))
        
        # Column C: Location
        ws.cell(row=row, column=3, value=job.get('location', 'N/A'))
        
        # Column D: Salary
        usd_salary = job.get('usd_equivalent')
        if usd_salary:
            ws.cell(row=row, column=4, value=f"${usd_salary:,.0f}")
        else:
            ws.cell(row=row, column=4, value=job.get('local_salary', 'N/A'))
        
        # Column E: Selected Resume (NEW)
        ws.cell(row=row, column=5, value=job.get('selected_resume', 'Not Available'))
        
        # Column F: Match Score (Enhanced)
        match_score = job.get('enhanced_match_score', job.get('match_score', 0))
        ws.cell(row=row, column=6, value=f"{match_score:.1f}%")
        
        # Column G: Selection Rationale (NEW)
        ws.cell(row=row, column=7, value=job.get('selection_rationale', ''))
        
        # Column H: Enhancement 1 (NEW)
        ws.cell(row=row, column=8, value=job.get('enhancement_1', ''))
        
        # Column I: Enhancement 2 (NEW)
        ws.cell(row=row, column=9, value=job.get('enhancement_2', ''))
        
        # Column J: Enhancement 3 (NEW)
        ws.cell(row=row, column=10, value=job.get('enhancement_3', ''))
        
        # Column K: Job Source
        ws.cell(row=row, column=11, value=job.get('source_site', 'N/A'))
        
        # Column L: Date Posted
        ws.cell(row=row, column=12, value=job.get('posted_date', 'N/A'))
    
    def _add_original_job_row(self, ws, row: int, job: Dict):
        """Add job row with original column structure.
        
        Args:
            ws: Excel worksheet
            row: Row number
            job: Job dictionary
        """
        ws.cell(row=row, column=1, value=job.get('title', 'N/A'))
        ws.cell(row=row, column=2, value=job.get('company', 'N/A'))
        ws.cell(row=row, column=3, value=job.get('location', 'N/A'))
        ws.cell(row=row, column=4, value=f"{job.get('match_score', 0):.2f}")
        
        # Salary formatting
        usd_salary = job.get('usd_equivalent')
        if usd_salary:
            ws.cell(row=row, column=5, value=f"${usd_salary:,.0f}")
        else:
            ws.cell(row=row, column=5, value=job.get('local_salary', 'N/A'))
        
        ws.cell(row=row, column=6, value="Yes" if job.get('visa_required') else "No")
        ws.cell(row=row, column=7, value=job.get('source_site', 'N/A'))
        ws.cell(row=row, column=8, value=job.get('posted_date', 'N/A'))
    
    def _apply_regional_styling(self, ws, region: str):
        """Apply regional color styling to worksheet.
        
        Args:
            ws: Excel worksheet
            region: Region name
        """
        # Get regional color
        region_color = self.regional_colors.get(region, 'FFFFFF')
        
        # Apply color to header rows
        for row in range(1, 8):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                    cell.fill = PatternFill(
                        start_color=region_color, 
                        end_color=region_color, 
                        fill_type="solid"
                    )
    
    def _get_top_company(self, jobs: List[Dict]) -> str:
        """Get the company with the most jobs in the list.
        
        Args:
            jobs: List of job dictionaries
            
        Returns:
            Company name with highest job count
        """
        if not jobs:
            return "N/A"
        
        company_counts = {}
        for job in jobs:
            company = job.get('company', 'Unknown')
            company_counts[company] = company_counts.get(company, 0) + 1
        
        return max(company_counts.items(), key=lambda x: x[1])[0]
    
    def _calculate_avg_salary(self, jobs: List[Dict]) -> Optional[float]:
        """Calculate average USD salary for jobs.
        
        Args:
            jobs: List of job dictionaries
            
        Returns:
            Average salary in USD or None
        """
        salaries = []
        for job in jobs:
            usd_salary = job.get('usd_equivalent')
            if usd_salary and isinstance(usd_salary, (int, float)):
                salaries.append(usd_salary)
        
        return sum(salaries) / len(salaries) if salaries else None
    
    def _get_region_emoji(self, region: str) -> str:
        """Get emoji for region.
        
        Args:
            region: Region name
            
        Returns:
            Emoji string
        """
        emojis = {
            'North America': '🇺🇸',
            'Western Europe': '🇪🇺',
            'Eastern Europe': '🌍',
            'East Asia': '🇯🇵',
            'Southeast Asia': '🌏',
            'South Asia': '🇮🇳',
            'South America': '🇧🇷',
            'Africa': '🌍',
            'Australia/Oceania': '🇦🇺',
            'Central America': '🌎',
            'Other': '🌐'
        }
        return emojis.get(region, '📍')
    
    def _adjust_column_widths_safe(self, ws):
        """Safely auto-adjust column widths for readability.
        
        Args:
            ws: Excel worksheet
        """
        try:
            # Use a simple approach to avoid MergedCell issues
            for col in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = chr(64 + col)  # A, B, C, etc.
                
                for row in range(1, min(ws.max_row + 1, 100)):  # Limit to first 100 rows
                    try:
                        cell = ws.cell(row=row, column=col)
                        if cell.value and not str(cell.coordinate) in str(ws.merged_cells):
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                    except:
                        continue
                
                # Set width with reasonable limits
                if max_length > 0:
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = max(adjusted_width, 10)
                else:
                    ws.column_dimensions[column_letter].width = 15
                    
        except Exception as e:
            logger.warning(f"Could not auto-adjust columns: {e}")
            # Fallback: set standard widths
            standard_widths = {
                'A': 25,  # Title
                'B': 20,  # Company
                'C': 25,  # Location
                'D': 12,  # Match Score
                'E': 15,  # Salary
                'F': 12,  # Visa
                'G': 15,  # Source
                'H': 12   # Date
            }
            
            for col_letter, width in standard_widths.items():
                try:
                    ws.column_dimensions[col_letter].width = width
                except:
                    pass
