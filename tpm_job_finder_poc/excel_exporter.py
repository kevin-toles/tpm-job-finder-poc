"""Excel export utilities with multi-resume intelligence support."""

import pandas as pd
from typing import Dict, List, Any, Optional
from pathlib import Path

from .models.resume_inventory import JobIntelligenceResult

# CORRECT column structure matching the exact specification
SPEC_COLUMNS = [
    'Job Title',            # Column A (Keep)
    'Company',              # Column B (Keep) 
    'Location',             # Column C (Keep)
    'Salary',               # Column D (Keep)
    'Selected Resume',      # Column E (NEW)
    'Match Score',          # Column F (Enhanced - was already there, now enhanced)
    'Selection Rationale',  # Column G (NEW)
    'Enhancement 1',        # Column H (NEW)
    'Enhancement 2',        # Column I (NEW)
    'Enhancement 3',        # Column J (NEW)
    'Job Source',           # Column K (Keep)
    'Date Posted',          # Column L (Keep)
]

# Column mappings exactly matching specification
SPEC_COLUMN_MAPPINGS = {
    'A': 'Job Title',           # Keep existing
    'B': 'Company',             # Keep existing
    'C': 'Location',            # Keep existing
    'D': 'Salary',              # Keep existing
    'E': 'Selected Resume',     # NEW: ai/ml_engineer.pdf
    'F': 'Match Score',         # Enhanced: 87.5% (was already there, now enhanced)
    'G': 'Selection Rationale', # NEW: Best ML keyword match
    'H': 'Enhancement 1',       # NEW: Led MLOps platform serving 10M+ predictions/day
    'I': 'Enhancement 2',       # NEW: Reduced model training costs by 60% via spot instances
    'J': 'Enhancement 3',       # NEW: Published 5 ML papers in top-tier conferences
    'K': 'Job Source',          # Keep existing
    'L': 'Date Posted',         # Keep existing
}

def export_to_excel(data, output_path):
    """Export data to Excel file following exact specification.
    
    Args:
        data: Data to export (dict, list, or DataFrame)
        output_path: Path for output Excel file
    """
    if isinstance(data, dict):
        df = pd.DataFrame([data])
    elif isinstance(data, list):
        df = pd.DataFrame(data) if data else pd.DataFrame()
    else:
        df = data
        
    # Ensure columns match exact specification (A-L structure)
    if not df.empty:
        # Reorder columns to match specification exactly
        available_columns = [col for col in SPEC_COLUMNS if col in df.columns]
        if available_columns:
            df = df[available_columns]
        
    df.to_excel(output_path, index=False, engine='openpyxl')

class SpecCompliantExcelExporter:
    """Excel exporter that follows the exact specification column structure."""
    
    def __init__(self):
        self.column_mappings = SPEC_COLUMN_MAPPINGS
        self.columns = SPEC_COLUMNS
    
    def export_with_multi_resume_intelligence(self, 
                                            jobs_data: List[Dict[str, Any]],
                                            intelligence_results: List[JobIntelligenceResult],
                                            output_path: str) -> None:
        """
        Export jobs with multi-resume intelligence following EXACT specification
        
        Columns A-D, K-L: Keep existing data
        Column F: Enhance existing Match Score with multi-resume logic
        Columns E, G, H, I, J: Add new multi-resume intelligence data
        
        Args:
            jobs_data: Original job data
            intelligence_results: Multi-resume intelligence results
            output_path: Path for output Excel file
        """
        # Create mapping of job_id to intelligence result
        intelligence_map = {result.job_id: result for result in intelligence_results}
        
        # Transform job data to match specification exactly
        enhanced_jobs = []
        for job in jobs_data:
            # Start with specification-compliant row structure
            spec_row = {
                'Job Title': job.get('title', job.get('Job Title', '')),              # Column A
                'Company': job.get('company', job.get('Company', '')),                # Column B  
                'Location': job.get('location', job.get('Location', '')),             # Column C
                'Salary': job.get('salary', job.get('Salary', '')),                   # Column D
                'Selected Resume': '',                                                 # Column E (NEW)
                'Match Score': job.get('match_score', job.get('Match Score', '0%')),  # Column F (Enhanced)
                'Selection Rationale': '',                                             # Column G (NEW)
                'Enhancement 1': '',                                                   # Column H (NEW)
                'Enhancement 2': '',                                                   # Column I (NEW)
                'Enhancement 3': '',                                                   # Column J (NEW)
                'Job Source': job.get('source', job.get('Job Source', '')),          # Column K
                'Date Posted': job.get('date_posted', job.get('Date Posted', '')),   # Column L
            }
            
            # Add intelligence data if available
            job_id = job.get('id', job.get('Job ID', ''))
            intelligence = intelligence_map.get(job_id)
            
            if intelligence:
                # Update with multi-resume intelligence
                spec_row.update({
                    'Selected Resume': intelligence.selected_resume.filename if intelligence.selected_resume else 'Not Available',
                    'Match Score': f"{intelligence.match_score:.1f}%" if intelligence.match_score else spec_row['Match Score'],
                    'Selection Rationale': intelligence.selection_rationale or 'Multi-resume processing completed',
                    'Enhancement 1': intelligence.enhancements[0].bullet_point if len(intelligence.enhancements) > 0 else 'None',
                    'Enhancement 2': intelligence.enhancements[1].bullet_point if len(intelligence.enhancements) > 1 else 'None',
                    'Enhancement 3': intelligence.enhancements[2].bullet_point if len(intelligence.enhancements) > 2 else 'None',
                })
            else:
                # Use fallback values for missing intelligence  
                spec_row.update({
                    'Selected Resume': 'Not Available',  # Avoid 'N/A' which becomes NaN
                    'Selection Rationale': 'Multi-resume processing not available',
                    'Enhancement 1': 'None',  # Use 'None' instead of empty string
                    'Enhancement 2': 'None',
                    'Enhancement 3': 'None',
                })
            
            enhanced_jobs.append(spec_row)
        
        # Export following exact specification
        df = pd.DataFrame(enhanced_jobs)
        
        # Ensure column order matches specification (A through L)
        df = df[SPEC_COLUMNS]
        
        # Export with regional tabs maintained
        self._export_with_regional_tabs(df, output_path)
    
    def _export_with_regional_tabs(self, df: pd.DataFrame, output_path: str) -> None:
        """Export with regional tab structure maintained"""
        # Regional mappings
        regional_mappings = {
            'North America': ['USA', 'Canada', 'Mexico', 'Remote', 'US', 'CA'],
            'Western Europe': ['UK', 'Germany', 'France', 'Netherlands', 'Spain', 'Italy'],
            'East Asia': ['China', 'Japan', 'Singapore', 'Hong Kong', 'South Korea'],
            'Oceania': ['Australia', 'New Zealand'],
            'Central America': ['Brazil', 'Argentina', 'Chile', 'Colombia'],
            'Middle East & Africa': ['UAE', 'Israel', 'South Africa', 'Egypt']
        }
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Process each region
            for region, countries in regional_mappings.items():
                # Filter jobs for this region
                region_filter = df['Location'].str.contains('|'.join(countries), case=False, na=False)
                region_df = df[region_filter]
                
                if not region_df.empty:
                    # Write to regional tab
                    region_df.to_excel(writer, sheet_name=region, index=False)
                    
                    # Apply formatting
                    self._format_regional_sheet(writer, region, region_df)
            
            # Add summary tab with all data
            df.to_excel(writer, sheet_name='All Regions', index=False)
            self._format_regional_sheet(writer, 'All Regions', df)
    
    def _format_regional_sheet(self, writer, sheet_name: str, df: pd.DataFrame) -> None:
        """Apply formatting to regional sheet following exact specification"""
        try:
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            worksheet = writer.sheets[sheet_name]
            
            # Set column widths exactly as specified
            column_widths = {
                'A': 25,  # Job Title
                'B': 20,  # Company  
                'C': 15,  # Location
                'D': 12,  # Salary
                'E': 18,  # Selected Resume (NEW)
                'F': 10,  # Match Score (Enhanced)
                'G': 15,  # Selection Rationale (NEW)
                'H': 40,  # Enhancement 1 (NEW)
                'I': 40,  # Enhancement 2 (NEW)
                'J': 40,  # Enhancement 3 (NEW)
                'K': 12,  # Job Source
                'L': 10,  # Date Posted
            }
            
            # Apply column widths
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Highlight NEW columns (E, G, H, I, J) with light blue background
            new_column_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            enhanced_column_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
            
            # Format headers
            header_font = Font(bold=True)
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                cell = worksheet[f'{col}1']
                cell.font = header_font
                
                # Highlight new columns
                if col in ['E', 'G', 'H', 'I', 'J']:
                    cell.fill = new_column_fill
                elif col == 'F':  # Enhanced column
                    cell.fill = enhanced_column_fill
            
            # Format data rows
            for row in range(2, len(df) + 2):
                for col in ['E', 'G', 'H', 'I', 'J']:
                    worksheet[f'{col}{row}'].fill = new_column_fill
                worksheet[f'F{row}'].fill = enhanced_column_fill
                    
        except ImportError:
            # openpyxl not available, skip formatting
            pass
        except Exception as e:
            # Log formatting error but don't fail export
            print(f"Warning: Could not apply formatting to {sheet_name}: {e}")


# Legacy alias for backward compatibility
EnhancedExcelExporter = SpecCompliantExcelExporter
